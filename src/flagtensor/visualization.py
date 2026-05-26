import csv
import os
from collections import defaultdict
from typing import Iterable

import ast
import matplotlib.pyplot as plt


def write_benchmark_csv(results: Iterable, output_path: str):
    results = list(results)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fieldnames = ["shape", "dtype", "mode", "latency", "latency_base", "speedup"]

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow(result.to_dict())

    mode_groups = defaultdict(list)
    for result in results:
        mode_groups[getattr(result, "mode", "operator")].append(result)

    root, ext = os.path.splitext(output_path)
    for mode, items in mode_groups.items():
        mode_output_path = f"{root}_{mode}{ext}"
        with open(mode_output_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for result in items:
                writer.writerow(result.to_dict())


def write_benchmark_xlsx(results: Iterable, output_path: str):
    """Standardized XLSX acceptance report.

    Produces a multi-sheet workbook:
      - Summary: two-level average speedup (by dtype, overall)
      - By Mode: per-mode sheet with all results
      - By Dtype: per-dtype summary with shape-level stats
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from collections import defaultdict

    results = list(results)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def _style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Summary sheet ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    _style_header(ws_summary)

    by_dtype_shape = defaultdict(lambda: defaultdict(list))
    for r in results:
        by_dtype_shape[r.dtype][r.shape].append(r.speedup)

    all_shape_avgs = []
    for dtype_name, shape_map in by_dtype_shape.items():
        shape_means = [sum(v) / len(v) for v in shape_map.values()]
        avg = sum(shape_means) / len(shape_means) if shape_means else 0.0
        ws_summary.append([f"Avg Speedup ({dtype_name})", f"{avg:.4f}x"])
        all_shape_avgs.extend(shape_means)

    overall = sum(all_shape_avgs) / len(all_shape_avgs) if all_shape_avgs else 0.0
    ws_summary.append(["Overall Avg Speedup", f"{overall:.4f}x"])
    ws_summary.append(["Total Results", len(results)])

    # ── By Mode sheets ──
    by_mode = defaultdict(list)
    for r in results:
        by_mode[getattr(r, "mode", "operator")].append(r)

    for mode_name, items in sorted(by_mode.items()):
        ws = wb.create_sheet(title=f"Mode-{mode_name}")
        ws.append(["Shape", "Dtype", "Triton (ms)", "Baseline (ms)", "Speedup"])
        _style_header(ws)
        for r in sorted(items, key=lambda x: (x.dtype, x.shape)):
            ws.append([
                str(r.shape),
                r.dtype,
                round(r.latency, 6) if r.latency else None,
                round(r.latency_base, 6) if r.latency_base else None,
                round(r.speedup, 4) if r.speedup else None,
            ])

    # ── By Dtype sheets ──
    by_dtype = defaultdict(list)
    for r in results:
        by_dtype[getattr(r, "dtype", "float32")].append(r)

    for dtype_name, items in sorted(by_dtype.items()):
        short = dtype_name.split(".")[-1] if isinstance(dtype_name, str) else str(dtype_name)
        ws = wb.create_sheet(title=f"Dtype-{short}")
        ws.append(["Shape", "Mode", "Triton (ms)", "Baseline (ms)", "Speedup"])
        _style_header(ws)
        for r in sorted(items, key=lambda x: (x.shape, getattr(x, "mode", "operator"))):
            ws.append([
                str(r.shape),
                getattr(r, "mode", "operator"),
                round(r.latency, 6) if r.latency else None,
                round(r.latency_base, 6) if r.latency_base else None,
                round(r.speedup, 4) if r.speedup else None,
            ])

    wb.save(output_path)


def plot_latency_and_speedup(results: Iterable, output_dir: str, op_name: str):
    os.makedirs(output_dir, exist_ok=True)
    grouped = defaultdict(list)
    for result in results:
        grouped[(result.dtype, getattr(result, "mode", "operator"))].append(result)

    for (dtype, mode), items in grouped.items():
        items = sorted(
            items,
            key=lambda item: ast.literal_eval(item.shape)[0] if isinstance(item.shape, str) else item.shape[0],
        )
        sizes = [ast.literal_eval(item.shape)[0] if isinstance(item.shape, str) else item.shape[0] for item in items]
        dtype_label = dtype.split(".")[-1] if isinstance(dtype, str) else str(dtype)
        triton_latency = [item.latency for item in items]
        cutensor_latency = [item.latency_base for item in items]
        speedup = [item.speedup for item in items]

        plt.figure(figsize=(8, 5))
        plt.plot(sizes, triton_latency, marker="o", label="Triton")
        plt.plot(sizes, cutensor_latency, marker="o", label="cuTensor")
        plt.xscale("log", base=2)
        plt.xlabel("Tensor size")
        plt.ylabel("Latency (ms)")
        plt.title(f"{op_name} Latency Comparison ({dtype}, {mode})")
        plt.legend()
        plt.grid(True, alpha=0.3)
        latency_path = os.path.join(output_dir, f"{dtype_label}_{mode}_latency.png")
        plt.tight_layout()
        plt.savefig(latency_path, dpi=200)
        plt.close()

        plt.figure(figsize=(8, 5))
        plt.plot(sizes, speedup, marker="o", label="Triton vs cuTensor")
        plt.axhline(1.0, linestyle="--", color="gray", linewidth=1)
        plt.xscale("log", base=2)
        plt.xlabel("Tensor size")
        plt.ylabel("Speedup (cuTensor / Triton)")
        plt.title(f"{op_name} Speedup Comparison ({dtype}, {mode})")
        plt.legend()
        plt.grid(True, alpha=0.3)
        speedup_path = os.path.join(output_dir, f"{dtype_label}_{mode}_speedup.png")
        plt.tight_layout()
        plt.savefig(speedup_path, dpi=200)
        plt.close()
